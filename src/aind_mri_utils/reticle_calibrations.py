import numpy as np
from scipy import optimize as opt
from scipy.spatial.transform import Rotation
from . import rotations
from . import utils as ut

from openpyxl import load_workbook


def extract_calibration_metadata(ws):
    rowiter = ws.iter_rows(min_row=1, max_row=2, values_only=True)
    colname_lookup = {k: i for i, k in enumerate(next(rowiter))}
    metadata_values = next(rowiter)
    global_factor = metadata_values[colname_lookup["GlobalFactor"]]
    global_rotation_degrees = metadata_values[
        colname_lookup["GlobalRotationDegrees"]
    ]
    manipulator_factor = metadata_values[colname_lookup["ManipulatorFactor"]]
    reticle_name = metadata_values[colname_lookup["Reticule"]]
    offset_x_pos = colname_lookup["GlobalOffsetX"]
    global_offset = np.array(
        metadata_values[offset_x_pos : offset_x_pos + 3], dtype=float
    )
    return (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    )


def extract_calibration_pairs(ws):
    pairs_by_probe = dict()
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        probe_name = row[0]
        if probe_name is None:
            continue
        reticle_pt = np.array(row[1:4])
        probe_pt = np.array(row[4:7])
        if probe_name not in pairs_by_probe:
            pairs_by_probe[probe_name] = []
        pairs_by_probe[probe_name].append((reticle_pt, probe_pt))
    return pairs_by_probe


def _combine_pairs(list_of_pairs):
    global_pts, manipulator_pts = [np.vstack(x) for x in zip(*list_of_pairs)]
    return global_pts, manipulator_pts


def _apply_metadata_to_pair_mats(
    global_pts,
    manipulator_pts,
    global_factor,
    global_rotation_degrees,
    global_offset,
    manipulator_factor,
):
    if global_rotation_degrees != 0:
        rotmat = (
            Rotation.from_euler("z", global_rotation_degrees, degrees=True)
            .as_matrix()
            .squeeze()
        )
        # Transposed because points are row vectors
        global_pts = global_pts @ rotmat.T
    global_pts = global_pts * global_factor + global_offset
    manipulator_pts = manipulator_pts * manipulator_factor
    return global_pts, manipulator_pts


def _apply_metadata_to_pair_lists(
    list_of_pairs,
    global_factor,
    global_rotation_degrees,
    global_offset,
    manipulator_factor,
):
    global_pts, manipulator_pts = _combine_pairs(list_of_pairs)
    return _apply_metadata_to_pair_mats(
        global_pts,
        manipulator_pts,
        global_factor,
        global_rotation_degrees,
        global_offset,
        manipulator_factor,
    )


def read_reticle_calibration(
    filename, points_sheet_name="points", metadata_sheet_name="metadata"
):
    wb = load_workbook(filename, read_only=True, data_only=True)
    if points_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {points_sheet_name} not found in {filename}")
    if metadata_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {metadata_sheet_name} not found in {filename}"
        )
    (
        global_factor,
        global_rotation_degrees,
        manipulator_factor,
        global_offset,
        reticle_name,
    ) = extract_calibration_metadata(wb[metadata_sheet_name])
    pairs_by_probe = extract_calibration_pairs(wb["points"])
    adjusted_pairs_by_probe = {
        k: _apply_metadata_to_pair_lists(
            v,
            global_factor,
            global_rotation_degrees,
            global_offset,
            manipulator_factor,
        )
        for k, v in pairs_by_probe.items()
    }
    return (
        adjusted_pairs_by_probe,
        global_offset,
        global_rotation_degrees,
        reticle_name,
    )


def _unpack_theta(theta):
    R = rotations.combine_angles(*theta[0:3])
    offset = theta[3:6]
    return R, offset


def fit_rotation_params(reticle_pts, probe_pts, legacy_outputs=False):
    R_homog = np.eye(4)
    reticle_pts_homog = ut.prepare_data_for_homogeneous_transform(reticle_pts)
    transformed_pts_homog = np.empty_like(reticle_pts_homog)

    def fun(theta):
        R_homog[0:3, 0:3] = rotations.combine_angles(*theta[0:3])
        R_homog[0:3, 3] = theta[3:6]  # translation
        np.matmul(reticle_pts_homog, R_homog.T, out=transformed_pts_homog)
        residuals = (transformed_pts_homog[:, 0:3] - probe_pts).flatten()
        return residuals

    theta0 = np.zeros(6)
    res = opt.least_squares(fun, theta0)
    R, translation = _unpack_theta(res.x)
    if legacy_outputs:
        # last version had translation in global frame
        #
        # All of the transposes are confusing here: this is the inverse of the
        # rotation matrix, accounting for numpy being row-major, and
        # data points being row vectors
        #
        # Also the last version found the tranpose of the rotation matrix
        # for some reason the application of the rotation matrix without
        # tranpose and was consistent if not correct
        translation = translation @ R  # Not R.T!
        return translation, R.T  # Not R!
    return R, translation


def apply_rotate_translate(pts, R, translation):
    R_homog = ut.make_homogeneous_transform(R, translation)
    pts_homog = ut.prepare_data_for_homogeneous_transform(pts)
    # Transposed because points are assumed to be row vectors
    transformed_pts_homog = pts_homog @ R_homog.T
    return ut.extract_data_for_homogeneous_transform(transformed_pts_homog)


def inverse_rotate_translate(R, translation):
    tinv = -translation @ R
    return R.T, tinv


def transform_reticle_to_probe(reticle_pts, R, translation):
    """
    Transform reticle points to probe points using rotation and translation.

    Parameters
    ----------
    probe_pts : np.array(N,3)
        Probe points to transform.
    R : np.array(3,3)
        Rotation matrix.
    translation : np.array(3,)
        Translation vector.

    Returns
    -------
    np.array(N,3)
        Transformed points.
    """
    return apply_rotate_translate(reticle_pts, R, translation)


def transform_probe_to_reticle(probe_pts, R, translation):
    """
    Transform probe points to reticle points using rotation and translation.

    Parameters
    ----------
    probe_pts : np.array(N,3)
        Probe points to transform.
    R : np.array(3,3)
        Rotation matrix.
    translation : np.array(3,)
        Translation vector.

    Returns
    -------
    np.array(N,3)
        Transformed points.
    """
    Rinv, tinv = inverse_rotate_translate(R, translation)
    return apply_rotate_translate(probe_pts, Rinv, tinv)
